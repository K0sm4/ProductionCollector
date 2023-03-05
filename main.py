import bs4
import requests
import openpyxl

def extract_data_from_tim_site(url):
    response = requests.get(url)

    # Load the page
    soup = bs4.BeautifulSoup(response.content, 'html.parser')
    product_details = soup.find_all('span', class_='base-text-clamp')
    product_description = soup.find_all('div', class_='product-header__text-wrapper')

    wb = openpyxl.load_workbook('listofcommercialitems.xlsx')
    max_row = wb.active.max_row

    # Function to add the next number
    item_number_cell = wb.active.cell(max_row, 1)
    next_item_number = int(item_number_cell.value) + 1
    item_number_cell = wb.active.cell(max_row + 1, 1)
    item_number_cell.value = next_item_number

    # Manufacturer
    manufacturer_cell = wb.active.cell(max_row + 1, 2)
    manufacturer_cell.value = str(product_details[0].text)

    # Model
    model_cell = wb.active.cell(max_row + 1, 3)
    model_cell.value = str(product_details[2].text)

    # Manufacturer number
    manufacturer_number_cell = wb.active.cell(max_row + 1, 4)
    manufacturer_number_cell.value = str(product_details[1].text)

    # Description
    description_cell = wb.active.cell(max_row + 1, 5)
    description_cell.value = str(product_description[0].text)

    wb.save('listofcommercialitems.xlsx')
    print("Data added successfully")

def extract_data_from_tme_site(url):
    response = requests.get(url)

    # Load the page
    soup = bs4.BeautifulSoup(response.content, 'html.parser')
    product_details = soup.find_all('span', class_='c-pip__symbol-value')
    product_description = soup.find_all('span', class_='c-pip__specification-parameter-value')
    manufacturers_designation = soup.find_all('span', class_='c-pip__symbol-value')
    description_details = soup.find_all('h2', class_='c-pip__h2')

    wb = openpyxl.load_workbook('listofcommercialitems.xlsx')
    max_row = wb.active.max_row

    # Function to add the next number
    item_number_cell = wb.active.cell(max_row, 1)
    next_item_number = int(item_number_cell.value) + 1
    item_number_cell = wb.active.cell(max_row + 1, 1)
    item_number_cell.value = next_item_number

    # Manufacturer
    manufacturer_cell = wb.active.cell(max_row + 1, 2)
    manufacturer_cell.value = str(product_description[0].text)

    # Model
    model_cell = wb.active.cell(max_row + 1, 3)
    model_cell.value = str(product_details[1].text)

    # Manufacturer number
    manufacturer_number_cell = wb.active.cell(max_row + 1, 4)
    manufacturer_number_cell.value = str(manufacturers_designation[0].text)

    # Description
    description_cell = wb.active.cell(max_row + 1, 5)
    description_cell.value = str(description_details[0].text)

    wb.save('listofcommercialitems.xlsx')
    print("Data added successfully")

# Get the URL from the user
url = input('Enter the URL:')
response = requests.get(url)

# Check if the connection to the internet is successful
if response.status_code == 200:
    content = response.text
    if 'tim.pl' in content:
        print('Data will be extracted from tim.pl site')
        extract_data_from_tim_site(url)
    elif 'tme.pl' in content:
        print('Data will be extracted from tme.pl site')
        extract_data_from_tme_site(url)
else:
        print('no internet connection')




